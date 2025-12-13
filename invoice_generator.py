import openpyxl
import csv
from pathlib import Path
from datetime import datetime

def generate_invoice(template_path, customer_data, output_path):
    """
    テンプレートから請求書を生成
    """
    # テンプレート読み込み
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    
    # データ書き込み
    ws['A4'] = customer_data['会社名']  # 会社名
    ws['B6'] = customer_data['件名']    # 件名
    ws['B10'] = f"¥{customer_data['金額']:,}"  # 合計金額
    ws['B14'] = f"¥{customer_data['金額']:,}"  # 内訳金額
    ws['B16'] = datetime.now().strftime('%Y年%m月%d日')  # 発行日
    
    # 保存
    wb.save(output_path)
    print(f"✅ 生成完了: {output_path}")

def batch_generate(template_path, csv_path, output_folder):
    """
    CSVから一括生成
    """
    print("=== 請求書一括生成開始 ===\n")
    
    # output_folderフォルダ作成
    Path(output_folder).mkdir(exist_ok=True)
    
    # CSVから顧客データ読み込み
    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        customers = list(reader)
    
    print(f"顧客数: {len(customers)}\n")
    
    # 各顧客の請求書を生成
    for idx, customer in enumerate(customers, 1):
        # 金額を整数に変換
        customer['金額'] = int(customer['金額'])
        
        # 出力ファイル名
        output_path = f"{output_folder}/請求書_{customer['会社名']}.xlsx"
        
        # 請求書生成
        generate_invoice(template_path, customer, output_path)
    
    print(f"\n=== 完了: {len(customers)}件の請求書を生成しました ===")

if __name__ == "__main__":
    batch_generate(
        template_path="template.xlsx",
        csv_path="input_data/customers.csv",
        output_folder="output_invoices"
    )
    