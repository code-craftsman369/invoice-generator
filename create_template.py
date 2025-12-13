import openpyxl
from openpyxl.styles import Font, Alignment

# 新規ワークブック作成
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "請求書"

# タイトル
ws['A1'] = '請　求　書'
ws['A1'].font = Font(size=20, bold=True)
ws['A1'].alignment = Alignment(horizontal='center')
ws.merge_cells('A1:D1')

# 宛先
ws['A3'] = '御中'
ws['A3'].font = Font(size=12)
ws['A4'] = '株式会社〇〇〇〇'  # ここに会社名が入る

# 件名
ws['A6'] = '件名:'
ws['B6'] = ''  # ここに件名が入る

# 本文
ws['A8'] = '下記の通りご請求申し上げます。'

# 合計金額
ws['A10'] = '合計金額'
ws['A10'].font = Font(bold=True)
ws['B10'] = ''  # ここに金額が入る
ws['C10'] = '円'

# 内訳
ws['A12'] = '内訳'
ws['A12'].font = Font(bold=True)
ws['A13'] = '項目'
ws['B13'] = '金額'
ws['A13'].font = Font(bold=True)
ws['B13'].font = Font(bold=True)

ws['A14'] = 'システム開発費'
ws['B14'] = ''  # ここに金額が入る

# 発行日
ws['A16'] = '発行日:'
ws['B16'] = ''  # ここに日付が入る

# 発行者
ws['A17'] = '株式会社△△△△'
ws['A17'].font = Font(bold=True)

# 列幅調整
ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 10

# 保存
wb.save('template.xlsx')
print("✅ テンプレート作成完了: template.xlsx")